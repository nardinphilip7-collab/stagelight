# qrcode_tracker/views.py
from django.shortcuts import render, redirect, HttpResponse
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import AuthenticationForm
import pandas as pd
import qrcode
from PIL import Image, ImageDraw, ImageFont
import uuid
from io import BytesIO
from .models import Attendee, Attendance, Configuration, UserProfile
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl

# Custom decorators for role-based access with permissions
def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            print("User not authenticated")
            return redirect('login')
        if not hasattr(request.user, 'userprofile'):
            print(f"No UserProfile for user: {request.user.username}")
            return render(request, 'qrcode_tracker/403.html', status=403)
        profile = request.user.userprofile
        if profile.role.lower() != 'admin':
            print(f"Invalid role for user {request.user.username}: {profile.role}")
            return render(request, 'qrcode_tracker/403.html', status=403)
        return view_func(request, *args, **kwargs)
    return wrapper

def admin_or_organizer_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            print("User not authenticated")
            return redirect('login')
        if not hasattr(request.user, 'userprofile'):
            print(f"No UserProfile for user: {request.user.username}")
            return render(request, 'qrcode_tracker/403.html', status=403)
        profile = request.user.userprofile
        if profile.role.lower() not in ['admin', 'organizer']:
            print(f"Invalid role for user {request.user.username}: {profile.role}")
            return render(request, 'qrcode_tracker/403.html', status=403)
        return view_func(request, *args, **kwargs)
    return wrapper

def permission_required(permission_name):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            if not hasattr(request.user, 'userprofile'):
                return render(request, 'qrcode_tracker/403.html', status=403)
            profile = request.user.userprofile
            if not getattr(profile, permission_name, False):
                print(f"User {request.user.username} lacks permission: {permission_name}")
                return render(request, 'qrcode_tracker/403.html', status=403)
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('upload_excel')
        else:
            return render(request, 'qrcode_tracker/login.html', {'form': form, 'error': 'Invalid username or password'})
    else:
        form = AuthenticationForm()
    return render(request, 'qrcode_tracker/login.html', {'form': form})

def configure_app(request):
    config, created = Configuration.objects.get_or_create(id=1)  # Singleton config
    
    if request.method == 'POST':
        app_name = request.POST.get('app_name')
        excel_columns = request.POST.getlist('excel_columns')
        qr_layout = request.POST.getlist('qr_layout')
        
        if not excel_columns or len(excel_columns) < 1:
            return render(request, 'qrcode_tracker/configure.html', {'config': config, 'error': 'At least one Excel column is required.'})
        
        config.app_name = app_name
        config.excel_columns = excel_columns
        config.qr_layout = qr_layout
        config.is_configured = True
        config.save()
        return redirect('upload_excel')
    
    return render(request, 'qrcode_tracker/configure.html', {'config': config})

def redirect_if_not_configured(view_func):
    def wrapper(request, *args, **kwargs):
        config = Configuration.objects.first()
        if not config or not config.is_configured:
            return redirect('configure_app')
        return view_func(request, *args, **kwargs)
    return wrapper

@login_required
@admin_required
@permission_required('can_upload_excel')
@redirect_if_not_configured
def upload_excel(request):
    config = Configuration.objects.first()
    if request.method == 'POST' and 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        try:
            df = pd.read_excel(excel_file)
            required_columns = config.excel_columns
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return render(request, 'qrcode_tracker/upload.html', {'error': f'Missing required columns: {", ".join(missing_columns)}', 'config': config})
            
            duplicates = []
            for index, row in df.iterrows():
                unique_values = [row[col] for col in required_columns]
                query = {f'{col}': val for col, val in zip(required_columns, unique_values) if col in ['name', 'job_title']}
                extra_data = {col.lower(): row[col] for col in required_columns if col not in ['name', 'job_title']}
                if not Attendee.objects.filter(**query).exists():
                    attendee_data = {
                        'name': row.get('name', ''),
                        'job_title': row.get('job_title', ''),
                        'extra_data': extra_data if extra_data else None,
                        'qr_code': str(uuid.uuid4())
                    }
                    Attendee.objects.create(**attendee_data)
                else:
                    duplicates.append(str(unique_values))
            if duplicates:
                error_msg = "The following entries were skipped as they already exist: " + ", ".join(duplicates)
                return render(request, 'qrcode_tracker/upload.html', {'error': error_msg, 'config': config})
            return redirect('attendee_list')
        except Exception as e:
            return render(request, 'qrcode_tracker/upload.html', {'error': str(e), 'config': config})
    elif request.method == 'GET' and 'download_template' in request.GET:
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Attendance Template"

        for col_idx, column_name in enumerate(config.excel_columns, 1):
            worksheet.cell(row=1, column=col_idx, value=column_name)

        worksheet.cell(row=2, column=1, value="John Doe")
        worksheet.cell(row=2, column=2, value="Engineer")
        if len(config.excel_columns) > 2:
            worksheet.cell(row=2, column=3, value="john@example.com")

        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="attendance_template_{config.app_name}.xlsx"'
        return response

    return render(request, 'qrcode_tracker/upload.html', {'config': config})

@login_required
@admin_required
@permission_required('can_generate_qr_stickers')
@redirect_if_not_configured
def generate_qr_stickers(request):
    config = Configuration.objects.first()
    attendees = Attendee.objects.all()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="qr_stickers.pdf"'
    
    c = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    
    for attendee in attendees:
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(attendee.qr_code)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        sticker = Image.new('RGB', (200, 250), 'white')
        draw = ImageDraw.Draw(sticker)
        
        sticker.paste(qr_img.resize((150, 150)), (25, 20))
        
        try:
            font = ImageFont.truetype("arial.ttf", 16)
        except:
            font = ImageFont.load_default()
            
        y_offset = 180
        for field in config.qr_layout:
            value = attendee.get_dynamic_field(field) if field.lower() not in ['name', 'job_title'] else getattr(attendee, field.lower(), '')
            if value:
                draw.text((10, y_offset), f"{field}: {value}", font=font, fill='black')
                y_offset += 20
        
        sticker.save("temp.png")
        c.drawImage("temp.png", 50, height - 300)
        c.showPage()
    
    c.save()
    return response

@login_required
@admin_or_organizer_required
@permission_required('can_scan_qr')
@redirect_if_not_configured
def scan_qr(request, code):
    config = Configuration.objects.first()
    try:
        attendee = Attendee.objects.get(qr_code=code)
        second_field = config.excel_columns[1] if len(config.excel_columns) > 1 else config.excel_columns[0]
        second_value = attendee.get_dynamic_field(second_field) if second_field.lower() not in ['name', 'job_title'] else getattr(attendee, second_field.lower(), '')
        if Attendance.objects.filter(attendee=attendee).exists():
            return JsonResponse({
                'status': 'already_scanned',
                'message': f'{attendee.name} ({second_value}) has already attended.'
            })
        else:
            Attendance.objects.create(attendee=attendee)
            return JsonResponse({
                'status': 'success',
                'message': f'Successfully recorded attendance for {attendee.name} ({second_value}).'
            })
    except Attendee.DoesNotExist:
        return JsonResponse({
            'status': 'not_found',
            'message': 'QR code not found.'
        })

@login_required
@admin_or_organizer_required
@permission_required('can_view_attendance')
@redirect_if_not_configured
def attendance_list(request):
    config = Configuration.objects.first()
    attendances = Attendance.objects.all().order_by('-timestamp')
    return render(request, 'qrcode_tracker/attendance_list.html', {
        'attendances': attendances,
        'config': config,
        'user_role': request.user.userprofile.role if hasattr(request.user, 'userprofile') else None
    })

@login_required
@admin_required
@permission_required('can_view_attendees')
@redirect_if_not_configured
def attendee_list(request):
    config = Configuration.objects.first()
    attendees = Attendee.objects.all()
    for attendee in attendees:
        attendee.qr_code_image = attendee.get_qr_code_image()
    return render(request, 'qrcode_tracker/attendee_list.html', {
        'attendees': attendees,
        'config': config,
        'columns': config.excel_columns,
        'user_role': request.user.userprofile.role if hasattr(request.user, 'userprofile') else None
    })

@login_required
@admin_or_organizer_required
@permission_required('can_scan_qr')
@redirect_if_not_configured
def scan_page(request):
    config = Configuration.objects.first()
    attendances = Attendance.objects.all().order_by('-timestamp')
    return render(request, 'qrcode_tracker/scan.html', {
        'attendances': attendances,
        'config': config,
        'user_role': request.user.userprofile.role if hasattr(request.user, 'userprofile') else None
    })

@login_required
@admin_required
@permission_required('can_flush_attendees')
@redirect_if_not_configured
def flush_attendees(request):
    if request.method == 'POST':
        Attendee.objects.all().delete()
        return redirect('attendee_list')
    return redirect('attendee_list')

@login_required
@admin_required
@permission_required('can_flush_attendance')
@redirect_if_not_configured
def flush_attendance(request):
    if request.method == 'POST':
        Attendance.objects.all().delete()
        return redirect('attendance_list')
    return redirect('attendance_list')

@login_required
@admin_required
@permission_required('can_configure_app')
@redirect_if_not_configured
def configure_app(request):  # Re-added to ensure it’s decorated with permissions
    config, created = Configuration.objects.get_or_create(id=1)  # Singleton config
    
    if request.method == 'POST':
        app_name = request.POST.get('app_name')
        excel_columns = request.POST.getlist('excel_columns')
        qr_layout = request.POST.getlist('qr_layout')
        
        if not excel_columns or len(excel_columns) < 1:
            return render(request, 'qrcode_tracker/configure.html', {'config': config, 'error': 'At least one Excel column is required.'})
        
        config.app_name = app_name
        config.excel_columns = excel_columns
        config.qr_layout = qr_layout
        config.is_configured = True
        config.save()
        return redirect('upload_excel')
    
    return render(request, 'qrcode_tracker/configure.html', {'config': config})